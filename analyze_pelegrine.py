import openpyxl

wb = openpyxl.load_workbook('CustosIP03.2026-SIstemanovo.xlsx', data_only=True)

# === ABA 1: Centros de Custos ===
ws = wb['Centros de Custos']
print("=" * 80)
print("ABA: Centros de Custos")
print("=" * 80)

# Header
headers = []
for col in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=col).value
    headers.append(h)
    print(f"  Col {col}: {h}")

print(f"\nTotal linhas: {ws.max_row}")

# Sample first 10 data rows
print("\n--- Primeiras 10 linhas de dados ---")
for row in range(2, min(12, ws.max_row + 1)):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        vals.append(v)
    print(f"  Linha {row}: {vals}")

# Unique values in key columns
print("\n--- Valores únicos em colunas-chave ---")

# Cod. C.C (col that contains CC codes)
cc_col = None
for i, h in enumerate(headers):
    if h and 'C.C' in str(h):
        cc_col = i + 1
        break

if cc_col:
    cc_values = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=cc_col).value
        if v:
            cc_values.add(v)
    print(f"\nCod. C.C (col {cc_col}): {len(cc_values)} valores únicos")
    for v in sorted(cc_values):
        print(f"  - {v}")

# Check other unique columns
for col_idx, h in enumerate(headers):
    if h and any(k in str(h).upper() for k in ['CONTA', 'TIPO', 'GRUPO', 'CLASSE']):
        vals = set()
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx + 1).value
            if v:
                vals.add(v)
        print(f"\n{h} (col {col_idx + 1}): {len(vals)} valores únicos")
        for v in sorted(vals, key=str):
            print(f"  - {v}")

# Last rows
print(f"\n--- Últimas 5 linhas ---")
for row in range(max(2, ws.max_row - 4), ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        vals.append(v)
    print(f"  Linha {row}: {vals}")

print("\n")

# === ABA 2: Filtros ===
ws2 = wb['Filtros']
print("=" * 80)
print("ABA: Filtros")
print("=" * 80)

# Header
headers2 = []
for col in range(1, ws2.max_column + 1):
    h = ws2.cell(row=1, column=col).value
    headers2.append(h)
    print(f"  Col {col}: {h}")

print(f"\nTotal linhas: {ws2.max_row}")

# All data
print("\n--- Todos os dados ---")
for row in range(2, ws2.max_row + 1):
    vals = []
    for col in range(1, ws2.max_column + 1):
        v = ws2.cell(row=row, column=col).value
        vals.append(v)
    # Skip empty rows
    if any(v is not None for v in vals):
        print(f"  Linha {row}: {vals}")
