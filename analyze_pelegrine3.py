import openpyxl

wb = openpyxl.load_workbook('CustosIP03.2026-NovoSistema.xlsx', data_only=True)

# === ABA 1: Custos IP 03.2026 ===
ws = wb['Custos IP 03.2026']
print("=" * 80)
print("ABA: Custos IP 03.2026")
print("=" * 80)

# Header
headers = []
for col in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=col).value
    headers.append(h)
    print(f"  Col {col}: {h}")

print(f"\nTotal linhas: {ws.max_row}")

# Sample first 15 data rows
print("\n--- Primeiras 15 linhas de dados ---")
for row in range(2, min(17, ws.max_row + 1)):
    vals = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals[headers[col-1] or f'col{col}'] = v
    print(f"  Linha {row}: {vals}")

# Unique values in key columns
print("\n--- Valores únicos em colunas-chave ---")

for col_idx, h in enumerate(headers):
    if h is None:
        continue
    h_upper = str(h).upper()
    if any(k in h_upper for k in ['C.C', 'CONTA', 'GRUPO', 'CC', 'CS', 'EMPRESA', 'ESP']):
        vals = set()
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx + 1).value
            if v is not None:
                vals.add(v)
        print(f"\n{h} (col {col_idx + 1}): {len(vals)} valores únicos")
        for v in sorted(vals, key=str):
            print(f"  - {v}")

# Nome do Centro de Custo unique values
nome_col = 1
nomes = set()
for row in range(2, ws.max_row + 1):
    v = ws.cell(row=row, column=nome_col).value
    if v:
        nomes.add(str(v).strip())
print(f"\nNome do Centro de Custo (col 1): {len(nomes)} valores únicos")
for v in sorted(nomes):
    print(f"  - {v}")

# CC and CS columns (14 and 15)
print("\n--- Colunas CC (14) e CS (15) ---")
cc_vals = set()
cs_vals = set()
for row in range(2, ws.max_row + 1):
    cc = ws.cell(row=row, column=14).value
    cs = ws.cell(row=row, column=15).value
    if cc: cc_vals.add(cc)
    if cs: cs_vals.add(cs)
print(f"CC values ({len(cc_vals)}):")
for v in sorted(cc_vals, key=str):
    print(f"  {v}")
print(f"\nCS values ({len(cs_vals)}):")
for v in sorted(cs_vals, key=str):
    print(f"  {v}")

# Total by CC
print("\n--- Total por CC ---")
cc_totals = {}
for row in range(2, ws.max_row + 1):
    cc = ws.cell(row=row, column=14).value
    valor = ws.cell(row=row, column=12).value or 0
    if cc:
        cc_totals[cc] = cc_totals.get(cc, 0) + float(valor)
for cc in sorted(cc_totals.keys(), key=str):
    print(f"  {cc}: R$ {cc_totals[cc]:,.2f}")
print(f"  TOTAL: R$ {sum(cc_totals.values()):,.2f}")

# Total by CS
print("\n--- Total por CS ---")
cs_totals = {}
for row in range(2, ws.max_row + 1):
    cs = ws.cell(row=row, column=15).value
    valor = ws.cell(row=row, column=12).value or 0
    if cs:
        cs_totals[cs] = cs_totals.get(cs, 0) + float(valor)
for cs in sorted(cs_totals.keys(), key=str):
    print(f"  {cs}: R$ {cs_totals[cs]:,.2f}")
print(f"  TOTAL: R$ {sum(cs_totals.values()):,.2f}")

# Top CCs by value
print("\n--- Top 30 Centros de Custo por Valor ---")
cc_nome_totals = {}
for row in range(2, ws.max_row + 1):
    cod_cc = ws.cell(row=row, column=3).value
    nome = ws.cell(row=row, column=1).value
    valor = ws.cell(row=row, column=12).value or 0
    if cod_cc:
        if cod_cc not in cc_nome_totals:
            cc_nome_totals[cod_cc] = {'nome': nome, 'total': 0}
        cc_nome_totals[cod_cc]['total'] += float(valor)
for cc in sorted(cc_nome_totals.keys(), key=lambda x: -cc_nome_totals[x]['total'])[:30]:
    info = cc_nome_totals[cc]
    print(f"  CC {cc:>5} | R$ {info['total']:>12,.2f} | {info['nome']}")

# Last 5 rows
print(f"\n--- Últimas 5 linhas ---")
for row in range(max(2, ws.max_row - 4), ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        vals.append(v)
    print(f"  Linha {row}: {vals}")

# === ABA 2: Filtros ===
ws2 = wb['Filtros']
print("\n" + "=" * 80)
print("ABA: Filtros - COMPLETA")
print("=" * 80)
headers2 = []
for col in range(1, ws2.max_column + 1):
    h = ws2.cell(row=1, column=col).value
    headers2.append(h)
print(f"Headers: {headers2}")
print()

for row in range(2, ws2.max_row + 1):
    vals = []
    for col in range(1, ws2.max_column + 1):
        v = ws2.cell(row=row, column=col).value
        vals.append(v)
    if any(v is not None for v in vals):
        print(f"  Linha {row}: {vals}")
