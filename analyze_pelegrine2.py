import openpyxl

wb = openpyxl.load_workbook('CustosIP03.2026-SIstemanovo.xlsx', data_only=True)

# === ABA 1: Centros de Custos - Grupos ===
ws = wb['Centros de Custos']
grupos = set()
nomes_cc = {}  # cod_cc -> nome
cc_grupos = {}  # cod_cc -> set of grupos
cc_valores = {}  # cod_cc -> total valor

for row in range(2, ws.max_row + 1):
    nome = ws.cell(row=row, column=1).value
    cod_cc = ws.cell(row=row, column=3).value
    grupo = ws.cell(row=row, column=4).value
    valor = ws.cell(row=row, column=12).value or 0
    cc_col = ws.cell(row=row, column=14).value
    cs_col = ws.cell(row=row, column=15).value
    
    if grupo:
        grupos.add(str(grupo).strip())
    if cod_cc:
        nomes_cc[cod_cc] = nome
        if cod_cc not in cc_grupos:
            cc_grupos[cod_cc] = set()
        if grupo:
            cc_grupos[cod_cc].add(str(grupo).strip())
        if cod_cc not in cc_valores:
            cc_valores[cod_cc] = 0
        cc_valores[cod_cc] += float(valor) if valor else 0

print("=" * 80)
print("GRUPOS ÚNICOS (col 4 - Grupo)")
print("=" * 80)
for g in sorted(grupos):
    print(f"  {g}")

print(f"\nTotal: {len(grupos)} grupos")

print("\n" + "=" * 80)
print("CENTROS DE CUSTO (Cod. C.C) COM NOME E TOTAL")
print("=" * 80)
for cc in sorted(cc_valores.keys(), key=lambda x: -cc_valores[x]):
    nome = nomes_cc.get(cc, '?')
    total = cc_valores[cc]
    if total > 0:
        print(f"  CC {cc:>5} | R$ {total:>12,.2f} | {nome}")

print(f"\nTotal geral: R$ {sum(cc_valores.values()):,.2f}")

# === CC and CS columns ===
print("\n" + "=" * 80)
print("COLUNAS CC e CS (col 14 e 15)")
print("=" * 80)
cc_vals = set()
cs_vals = set()
for row in range(2, ws.max_row + 1):
    cc = ws.cell(row=row, column=14).value
    cs = ws.cell(row=row, column=15).value
    if cc: cc_vals.add(cc)
    if cs: cs_vals.add(cs)

print(f"CC values ({len(cc_vals)}):")
for v in sorted(cc_vals):
    print(f"  {v}")
print(f"\nCS values ({len(cs_vals)}):")
for v in sorted(cs_vals):
    print(f"  {v}")

# === ABA 2: Filtros - Full dump ===
ws2 = wb['Filtros']
print("\n" + "=" * 80)
print("ABA FILTROS - COMPLETA")
print("=" * 80)
headers2 = []
for col in range(1, ws2.max_column + 1):
    h = ws2.cell(row=1, column=col).value
    headers2.append(h)
print(f"Headers: {headers2}")
print()

for row in range(2, ws2.max_row + 1):
    vals = []
    for col in range(1, ws2.max_column + 1):
        v = ws2.cell(row=row, column=col).value
        vals.append(v)
    if any(v is not None for v in vals):
        print(f"  {vals}")
