"""
Análise final da planilha CustosIP03.2026-NovoSistema.xlsx
Aplicando as regras de filtro definidas pelo usuário:
- Aba Filtros define quais CCs participam da apuração
- Empresa "Irmãos Pelegrine" → só despesas da IRMAOS PELEGRINE
- Empresa "Todas" → todas as despesas do CC
- Empresa "Todas (30%)" → 30% do valor total (CCs 1527 e 3143)
"""
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('CustosIP03.2026-NovoSistema.xlsx', data_only=True)

# === 1. PARSE ABA FILTROS ===
ws_filtros = wb['Filtros']
filtros = {}  # cc_code -> { empresa, descricao, status, agrupamento }
for row in range(2, ws_filtros.max_row + 1):
    cc = ws_filtros.cell(row=row, column=1).value
    if cc is None or cc == 'Legenda:':
        continue
    status = ws_filtros.cell(row=row, column=2).value
    agrup = ws_filtros.cell(row=row, column=3).value
    empresa = ws_filtros.cell(row=row, column=4).value
    desc = ws_filtros.cell(row=row, column=5).value
    filtros[int(cc)] = {
        'empresa': str(empresa).strip() if empresa else None,
        'descricao': str(desc).strip() if desc else '',
        'status': str(status).strip() if status else '',
        'agrupamento': agrup,
    }

print(f"=== ABA FILTROS: {len(filtros)} centros de custo ===")
print(f"  - 'Irmãos Pelegrine': {sum(1 for f in filtros.values() if f['empresa'] == 'Irmãos Pelegrine')}")
print(f"  - 'Todas': {sum(1 for f in filtros.values() if f['empresa'] == 'Todas')}")
print(f"  - 'Todas (30%)': {sum(1 for f in filtros.values() if f['empresa'] == 'Todas (30%)')}")
print(f"  - Com status 'S': {sum(1 for f in filtros.values() if f['status'] == 'S')}")
print(f"  - Com status 'NT': {sum(1 for f in filtros.values() if f['status'] == 'NT')}")

# CCs com 30%
ccs_30pct = [cc for cc, f in filtros.items() if f['empresa'] == 'Todas (30%)']
print(f"\n  CCs com 30%: {ccs_30pct}")
for cc in ccs_30pct:
    print(f"    CC {cc}: {filtros[cc]['descricao']}")

# === 2. PARSE ABA CUSTOS ===
ws = wb['Custos IP 03.2026']
itens = []
for row in range(2, ws.max_row + 1):
    nome_cc = ws.cell(row=row, column=1).value
    empresa = ws.cell(row=row, column=2).value
    cod_cc = ws.cell(row=row, column=3).value
    grupo = ws.cell(row=row, column=4).value
    subgrupo = ws.cell(row=row, column=5).value
    descricao = ws.cell(row=row, column=6).value
    qtd = ws.cell(row=row, column=7).value
    esp = ws.cell(row=row, column=8).value
    lanc_contabil = ws.cell(row=row, column=9).value
    data = ws.cell(row=row, column=10).value
    referencia = ws.cell(row=row, column=11).value
    valor = ws.cell(row=row, column=12).value or 0
    
    if cod_cc is not None:
        itens.append({
            'nome_cc': str(nome_cc).strip() if nome_cc else '',
            'empresa': str(empresa).strip() if empresa else '',
            'cod_cc': int(cod_cc),
            'grupo': str(grupo).strip() if grupo else '',
            'subgrupo': str(subgrupo).strip() if subgrupo else '',
            'descricao': str(descricao).strip() if descricao else '',
            'qtd': qtd,
            'esp': str(esp).strip() if esp else '',
            'lanc_contabil': lanc_contabil,
            'data': data,
            'referencia': str(referencia).strip() if referencia else '',
            'valor': float(valor),
        })

print(f"\n=== ABA CUSTOS: {len(itens)} itens totais ===")

# === 3. APLICAR FILTROS ===
itens_filtrados = []
itens_excluidos = 0
itens_sem_filtro = 0

for item in itens:
    cc = item['cod_cc']
    if cc not in filtros:
        itens_sem_filtro += 1
        continue
    
    filtro = filtros[cc]
    empresa_filtro = filtro['empresa']
    
    if empresa_filtro == 'Irmãos Pelegrine':
        if item['empresa'] == 'IRMAOS PELEGRINE':
            itens_filtrados.append({**item, 'fator': 1.0})
        else:
            itens_excluidos += 1
    elif empresa_filtro == 'Todas':
        itens_filtrados.append({**item, 'fator': 1.0})
    elif empresa_filtro == 'Todas (30%)':
        itens_filtrados.append({**item, 'fator': 0.30})
    else:
        itens_sem_filtro += 1

print(f"\n=== RESULTADO DO FILTRO ===")
print(f"  Itens que passaram no filtro: {len(itens_filtrados)}")
print(f"  Itens excluídos (empresa diferente): {itens_excluidos}")
print(f"  Itens sem CC no filtro: {itens_sem_filtro}")

# Valor total filtrado
valor_total = sum(i['valor'] * i['fator'] for i in itens_filtrados)
print(f"  Valor total filtrado: R$ {valor_total:,.2f}")

# === 4. ANÁLISE POR HIERARQUIA (Nome do Centro de Custo) ===
print(f"\n=== ANÁLISE POR CENTRO DE CUSTO (filtrado) ===")
por_cc = defaultdict(lambda: {'valor': 0, 'qtd_itens': 0, 'desc': '', 'empresa_filtro': ''})
for item in itens_filtrados:
    cc = item['cod_cc']
    por_cc[cc]['valor'] += item['valor'] * item['fator']
    por_cc[cc]['qtd_itens'] += 1
    por_cc[cc]['desc'] = filtros[cc]['descricao']
    por_cc[cc]['empresa_filtro'] = filtros[cc]['empresa']

# Ordenar por valor decrescente
for cc, data in sorted(por_cc.items(), key=lambda x: -x[1]['valor']):
    if data['valor'] > 0:
        print(f"  CC {cc:5d} | R$ {data['valor']:>12,.2f} | {data['qtd_itens']:4d} itens | [{data['empresa_filtro']}] | {data['desc'][:80]}")

# === 5. ANÁLISE POR GRUPO (tipo de despesa) ===
print(f"\n=== ANÁLISE POR GRUPO (filtrado) ===")
por_grupo = defaultdict(lambda: {'valor': 0, 'qtd_itens': 0})
for item in itens_filtrados:
    g = item['grupo']
    por_grupo[g]['valor'] += item['valor'] * item['fator']
    por_grupo[g]['qtd_itens'] += 1

for g, data in sorted(por_grupo.items(), key=lambda x: -x[1]['valor']):
    print(f"  {g:50s} | R$ {data['valor']:>12,.2f} | {data['qtd_itens']:4d} itens")

# === 6. CLASSIFICAÇÃO POR CONTA DE CUSTO (regras do usuário) ===
print(f"\n=== CLASSIFICAÇÃO POR CONTA DE CUSTO ===")
# Regras do usuário:
# Combustível: Óleo diesel, gasolina e álcool
# Lubrificantes: Óleos, graxas e fluidos hidráulicos
# Peças de Desgaste: Pneus, unhas, dentes, mandíbulas, telas, roletes, bits
# Outras Despesas: Fretes, serviços de terceiros, mão-de-obra, lavagem, recapagem, pintura
# Peças de Reposição / Itens de Consumo: Tudo que não se enquadra (residual)

def classificar_conta(grupo, descricao):
    grupo_upper = grupo.upper()
    desc_upper = descricao.upper() if descricao else ''
    
    # Combustível
    if 'COMBUSTIVEL' in grupo_upper or 'COMBUSTÍVEL' in grupo_upper:
        # Dentro de COMBUSTIVEL/LUBRIF, separar
        if any(x in desc_upper for x in ['DIESEL', 'GASOLINA', 'ALCOOL', 'ÁLCOOL', 'ETANOL']):
            return 'Combustível'
        elif any(x in desc_upper for x in ['OLEO', 'ÓLEO', 'GRAXA', 'FLUIDO', 'HIDRAULIC', 'HIDRÁULIC', 'LUBRIF']):
            return 'Lubrificantes'
        else:
            return 'Combustível'  # default para o grupo COMBUSTIVEL
    
    # Provisão Folha / Salários
    if 'PROVISAO FOLHA' in grupo_upper or 'PRO-LABORE' in grupo_upper:
        return 'Salários/Folha'
    
    # Salários a pagar
    if 'SALARIO A PAGAR' in grupo_upper or 'RESCISAO A PAGAR' in grupo_upper:
        return 'Salários/Folha'
    
    # Explosivos
    if 'EXPLOSIVO' in grupo_upper:
        return 'Explosivos'
    
    # Material de Perfuração
    if 'PERFURACAO' in grupo_upper or 'PERFURAÇÃO' in grupo_upper:
        return 'Material de Perfuração'
    
    # Pneus → Peças de Desgaste
    if 'PNEU' in grupo_upper:
        return 'Peças de Desgaste'
    
    # Revestimento/Mandíbula → Peças de Desgaste
    if 'REVESTIMENTO' in grupo_upper or 'MANDIBULA' in grupo_upper:
        return 'Peças de Desgaste'
    
    # Roletes → Peças de Desgaste
    if 'ROLETE' in grupo_upper:
        return 'Peças de Desgaste'
    
    # Correias → Peças de Desgaste
    if 'CORREIA' in grupo_upper:
        return 'Peças de Desgaste'
    
    # Serviços → Outras Despesas
    if 'SERVICO' in grupo_upper or 'SERVIÇO' in grupo_upper:
        return 'Outras Despesas'
    
    # Alimentação → Outras Despesas
    if 'ALIMENTACAO' in grupo_upper or 'ALIMENTAÇÃO' in grupo_upper:
        return 'Outras Despesas'
    
    # Material de Segurança → Outras Despesas
    if 'SEGURANCA' in grupo_upper or 'SEGURANÇA' in grupo_upper:
        return 'Outras Despesas'
    
    # Financeiro
    if 'EMPRESTIMO' in grupo_upper or 'TARIFA' in grupo_upper or 'TAXA' in grupo_upper:
        return 'Despesas Financeiras'
    
    # Ajuda de custo, doação, retirada de lucro
    if any(x in grupo_upper for x in ['AJUDA DE CUSTO', 'DOACAO', 'RETIRADA DE LUCRO', 'DAS A PAGAR']):
        return 'Outras Despesas'
    
    # Seguros
    if 'SEGURO' in grupo_upper:
        return 'Outras Despesas'
    
    # Tudo mais → Peças de Reposição / Itens de Consumo
    return 'Peças de Reposição / Itens de Consumo'

por_conta = defaultdict(lambda: {'valor': 0, 'qtd_itens': 0, 'grupos': set()})
for item in itens_filtrados:
    conta = classificar_conta(item['grupo'], item['descricao'])
    por_conta[conta]['valor'] += item['valor'] * item['fator']
    por_conta[conta]['qtd_itens'] += 1
    por_conta[conta]['grupos'].add(item['grupo'])

for conta, data in sorted(por_conta.items(), key=lambda x: -x[1]['valor']):
    print(f"\n  {conta}: R$ {data['valor']:,.2f} ({data['qtd_itens']} itens)")
    print(f"    Grupos: {', '.join(sorted(data['grupos']))}")

# === 7. IDENTIFICAR CCs QUE SÃO EQUIPAMENTOS vs SETORES ===
print(f"\n=== CCs: EQUIPAMENTOS vs SETORES ===")
equipamentos = []
setores = []
for cc, f in sorted(filtros.items()):
    desc = f['descricao']
    # Equipamentos geralmente têm "EQUIPAMENTOS" na hierarquia
    if 'EQUIPAMENTOS' in desc.upper() and any(x in desc.upper() for x in [
        'BRITADOR', 'REBRITADOR', 'PENEIRA', 'TRANSPORTADOR', 'CALHA',
        'CARRETA', 'COMPRESSOR', 'CARREGADEIRA', 'ESCAVADEIRA', 'EMPILHADEIRA',
        'CACAMBA', 'MUNCK', 'TRATOR', 'CARRO PIPA', 'FIAT', 'L200', 'MOTO',
        'UNO', 'TORNO', 'FRESA', 'FURADEIRA', 'SOLDA', 'LABORATORIO',
        'RK430', 'PC240', 'CX220', 'HYUNDAI', 'WA200', '621', '966',
        'TRACADA', 'BRANCA'
    ]):
        equipamentos.append((cc, desc))
    else:
        setores.append((cc, desc))

print(f"\n  EQUIPAMENTOS ({len(equipamentos)}):")
for cc, desc in equipamentos:
    v = por_cc.get(cc, {}).get('valor', 0)
    print(f"    CC {cc:5d} | R$ {v:>10,.2f} | {desc[:70]}")

print(f"\n  SETORES/OPERACIONAL ({len(setores)}):")
for cc, desc in setores:
    v = por_cc.get(cc, {}).get('valor', 0)
    print(f"    CC {cc:5d} | R$ {v:>10,.2f} | {desc[:70]}")

# === 8. RESUMO FINAL ===
print(f"\n{'='*80}")
print(f"RESUMO FINAL DA APURAÇÃO DE CUSTOS PELEGRINE - MARÇO/2026")
print(f"{'='*80}")
print(f"  Total de CCs na aba Filtros: {len(filtros)}")
print(f"  Total de itens filtrados: {len(itens_filtrados)}")
print(f"  Valor total das despesas (filtrado): R$ {valor_total:,.2f}")
print(f"  + Impostos (lançamento manual): R$ 138.224,86")
print(f"  = TOTAL GERAL ESTIMADO: R$ {valor_total + 138224.86:,.2f}")
print(f"  Equipamentos identificados: {len(equipamentos)}")
print(f"  Setores/Operacional: {len(setores)}")
